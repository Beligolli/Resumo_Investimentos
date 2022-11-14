# %%
import pandas as pd
import win32com.client as win32
import time
import requests
import json
import openpyxl
import pandas_datareader.data as web
from datetime import date, timedelta
from openpyxl import load_workbook

# %% [markdown]
# Puxar Valor Dolar

# %%
requisicao = requests.get(r'https://economia.awesomeapi.com.br/all/USD-BRL')
cotacao = requisicao.json()
cotacao_dolar = float(cotacao['USD']['bid'])
print(cotacao_dolar)

# %% [markdown]
# Puxar quantidade Dolar existente

# %% [markdown]
# Atualizar planilha com bolsa Down Jones

# %%
carteira_df = pd.read_excel(
    r'C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\Carteira_usa.xlsx')
today = date.today()
yesterday = today - timedelta(days = 1)
cotacoes_carteira = pd.DataFrame()

try:
    for ativo in carteira_df['Ativos']:
        cotacoes_carteira[ativo] = web.DataReader('{}'.format(ativo), data_source='yahoo', start=yesterday, end=yesterday)['Adj Close']
except:
        for ativo in carteira_df['Ativos']:
            cotacoes_carteira[ativo] = web.DataReader('{}'.format(ativo), data_source='yahoo', start=today, end=today)['Adj Close']

valor_investido = pd.DataFrame()
for ativo in carteira_df['Ativos']:
    valor_investido[ativo] = cotacoes_carteira[ativo] * carteira_df.loc[carteira_df['Ativos'] == ativo, 'Qtde'].values[0]

valor_investido['TOTAL'] = valor_investido[list(valor_investido.columns)].sum(axis=1)

valor_investido.to_excel(r'C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\Carteira_usa_dia.xlsx')

# %%
dolar_wb = load_workbook(r"C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\Carteira_usa_dia.xlsx", data_only=True)
sh = dolar_wb["Sheet1"]
dolartt= (sh["k2"].value)
print(dolartt)

# %% [markdown]
# Puxando montante em criptos

# %%
carteira_cripto_df = pd.read_excel(
    r'C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\Carteira_cripto.xlsx')
today = date.today()
yesterday = today - timedelta(days = 1)
cotacoes_carteira = pd.DataFrame()

for ativo in carteira_cripto_df['Ativos']:
    cotacoes_carteira[ativo] = web.DataReader('{}'.format(ativo), data_source='yahoo', start=today, end=today)['Adj Close']

valor_investido = pd.DataFrame()
for ativo in carteira_cripto_df['Ativos']:
    valor_investido[ativo] = cotacoes_carteira[ativo] * carteira_cripto_df.loc[carteira_cripto_df['Ativos'] == ativo, 'Qtde'].values[0]

valor_investido['TOTAL'] = valor_investido[list(valor_investido.columns)].sum(axis=1)

valor_investido.to_excel(r"C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\Carteira_criptos_dia.xlsx")

# %%
cripto_wb = load_workbook(r"C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\Carteira_criptos_dia.xlsx", data_only=True)
sh = cripto_wb["Sheet1"]
criptott= (sh["E2"].value)
print(criptott)

# %% [markdown]
# Abrindo Data Frames

# %% [markdown]
# Abrir Info.xls e alterar o valor do dolar cota��o e dolar montante total

# %%

wb = load_workbook(filename=r'C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\infos.xlsx')
ws = wb.worksheets[1]
ws['I4'] = cotacao_dolar
ws['I2'] = dolartt
ws['H2'] = criptott
wb.save(r'C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\infos.xlsx')


# %% [markdown]
# Excel file, open, refresh and close

# %%
# Start an instance of Excel
xlapp = win32.DispatchEx("Excel.Application")

# Open the workbook in said instance of Excel
wb = xlapp.workbooks.open(r'C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\infos.xlsx')

# Optional, e.g. if you want to debug
xlapp.Visible = True

# Refresh all data connections.
wb.RefreshAll()
wb.Save()

# Quit
xlapp.Quit()

# %%
# filenames
excel_names = [r"C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\base_dash.xlsx", r"C:\Users\belig\OneDrive\Python\MeuProjeto\Projetos\dash_investimentos\Resumo_Investimentos\infos.xlsx"]

# read them in
excels = [pd.ExcelFile(name) for name in excel_names]

# turn them into dataframes
frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]

# delete the first row for all frames except the first
# i.e. remove the header row -- assumes it's the first
frames[1:] = [df[1:] for df in frames[1:]]

# concatenate them..
combined = pd.concat(frames)

# write it out
combined.to_excel("base_dash.xlsx", header=False, index=False)


